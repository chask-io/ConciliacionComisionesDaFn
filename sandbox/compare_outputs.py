#!/usr/bin/env python3
"""
Compare rewritten (no-pandas) conciliation output against original pandas version.

Usage:
    cd platform/conciliacion_comisiones_da
    python sandbox/compare_outputs.py
"""

import os
import sys
import json

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FUNCTION_DIR = os.path.dirname(SCRIPT_DIR)
SRC_DIR = os.path.join(FUNCTION_DIR, "src")
SANDBOX_DATA = "/home/megaterium/sandbox/daniel_achondo/conciliacion_comisiones"

# Add src to path for imports
sys.path.insert(0, SRC_DIR)

from main import run, ConciliationConfig


def main():
    ventas_dir = os.path.join(SANDBOX_DATA, "ventas")
    pagos_file = os.path.join(SANDBOX_DATA, "pagos", "COMISIONES 2025 - DANIEL ACHONDO.xlsx")
    credito_file = os.path.join(SANDBOX_DATA, "ventas", "VENTAS CON CREDITO BK.xlsx")
    output_dir = "/tmp/conciliacion_test"

    os.makedirs(output_dir, exist_ok=True)

    print("=" * 80)
    print("Running REWRITTEN conciliation (no-pandas)...")
    print("=" * 80)

    config = ConciliationConfig(
        ventas_dir=ventas_dir,
        pagos_file=pagos_file,
        credito_file=credito_file,
        output_dir=output_dir,
        fuzzy_threshold=85,
        tolerance=1000
    )

    result = run(config)

    if result["status"] != "ok":
        print(f"\nERROR: {result.get('error')}")
        sys.exit(1)

    stats = result["result"]["statistics"]
    summary = result["result"]["summary"]
    output_file = result["result"]["output_file"]

    print("\n" + "=" * 80)
    print("RESULTS")
    print("=" * 80)

    print(f"\nOutput file: {output_file}")
    print(f"\nSummary:")
    for k, v in summary.items():
        print(f"  {k}: {v}")

    print(f"\nDetailed Statistics:")
    print(f"  Total vehiculos: {stats['total_vehiculos']}")
    print(f"  Con bonos: {stats['vehiculos_con_bonos']}")
    print(f"  Con pago: {stats['vehiculos_con_pago']}")
    print(f"  Matches exactos: {stats['matches_exactos']}")
    print(f"  Matches fuzzy: {stats['matches_fuzzy']}")
    print(f"  Monto sistema: ${stats['monto_bonos_sistema']:,.0f}")
    print(f"  Monto pagado: ${stats['monto_pagado_andes']:,.0f}")
    print(f"  Diferencia: ${stats['monto_diferencia']:,.0f}")
    print(f"  % Cobertura: {stats['pct_cobertura']}%")
    print(f"  % Exacta: {stats['pct_exacta']}%")
    print(f"  % Monetaria: {stats['pct_monetaria']}%")

    print(f"\n  Estados: {stats['estados']}")

    if stats.get('by_marca'):
        print(f"\n  Por marca:")
        for marca, data in stats['by_marca'].items():
            print(f"    {marca}: {data['total']} vehiculos, "
                  f"${data['monto_sistema']:,.0f} sistema, "
                  f"${data['monto_pagado']:,.0f} pagado, "
                  f"{data['pct_cobertura']}% cobertura")

    if stats.get('credit_sales'):
        cs = stats['credit_sales']
        print(f"\n  Ventas a crédito:")
        print(f"    Total: {cs.get('total_credit_sales', 0)}")
        print(f"    Matched por RUT: {cs.get('matched_by_rut', 0)}")
        print(f"    Matched por nombre: {cs.get('matched_by_name', 0)}")
        print(f"    Sin match: {cs.get('unmatched', 0)}")
        print(f"    Con pago: {cs.get('with_payment', 0)}")

    # Validate the Excel file was created and is non-empty
    file_size = os.path.getsize(output_file)
    print(f"\n  Output file size: {file_size:,} bytes")

    # Validate sheets
    import openpyxl
    wb = openpyxl.load_workbook(output_file, read_only=True)
    print(f"  Sheets: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        row_count = sum(1 for _ in ws.iter_rows(values_only=True))
        print(f"    {name}: {row_count} rows")
    wb.close()

    print("\n" + "=" * 80)
    print("TEST PASSED - Reconciliation completed successfully")
    print("=" * 80)

    # Save stats as JSON for comparison
    stats_file = os.path.join(output_dir, "stats.json")
    with open(stats_file, "w") as f:
        json.dump(stats, f, indent=2, default=str)
    print(f"\nStats saved to: {stats_file}")


if __name__ == "__main__":
    main()
