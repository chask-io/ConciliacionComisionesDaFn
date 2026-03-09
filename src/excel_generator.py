"""
Generador de reportes Excel para resultados de conciliación.

Genera reportes profesionales en Excel con:
- Hoja de resumen ejecutivo con KPIs
- Hoja de conciliación detallada por vehículo
- Hoja de pendientes de pago
- Hoja de coincidencias difusas para revisión
- Hoja de detalle de pagos de Andes
- Hoja de VINs sin venta registrada
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
from typing import Dict, List, Any, Optional
import os

try:
    from .utils import COLORES, MESES, format_currency, is_empty
except ImportError:
    from utils import COLORES, MESES, format_currency, is_empty

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Genera reportes Excel profesionales para resultados de conciliación."""

    def __init__(
        self,
        conciliacion: List[Dict],
        statistics: Dict[str, Any],
        pagos: List[Dict],
        unmatched_pagos: List[Dict],
        fuzzy_matches: List[Dict] = None,
        credit_status: List[Dict] = None
    ):
        self.conciliacion = conciliacion
        self.statistics = statistics
        self.pagos = pagos
        self.unmatched_pagos = unmatched_pagos
        self.fuzzy_matches = fuzzy_matches or []
        self.credit_status = credit_status

        self._setup_styles()

    def _setup_styles(self):
        """Configura estilos reutilizables."""
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.header_fill = PatternFill(start_color=COLORES['header'], end_color=COLORES['header'], fill_type='solid')
        self.header_fill_green = PatternFill(start_color=COLORES['header_green'], end_color=COLORES['header_green'], fill_type='solid')
        self.header_fill_red = PatternFill(start_color=COLORES['header_red'], end_color=COLORES['header_red'], fill_type='solid')
        self.header_fill_orange = PatternFill(start_color=COLORES['header_orange'], end_color=COLORES['header_orange'], fill_type='solid')
        self.fill_pagado = PatternFill(start_color=COLORES['pagado'], end_color=COLORES['pagado'], fill_type='solid')
        self.fill_pendiente = PatternFill(start_color=COLORES['pendiente'], end_color=COLORES['pendiente'], fill_type='solid')
        self.fill_parcial = PatternFill(start_color=COLORES['parcial'], end_color=COLORES['parcial'], fill_type='solid')
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.number_format = '#,##0'
        self.currency_format = '$#,##0'

    def generate_report(self, output_path: str):
        """Genera reporte Excel completo."""
        logger.info(f"Generando reporte Excel: {output_path}")

        wb = openpyxl.Workbook()

        datos_last_row = self._write_datos_sheet(wb)
        self._write_resumen_sheet(wb, datos_last_row)
        self._write_conciliacion_sheets_by_month(wb)
        self._write_pendientes_sheet(wb)
        self._write_fuzzy_sheet(wb)
        self._write_pagos_sheet(wb)
        self._write_sin_venta_sheet(wb)

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        wb.save(output_path)
        logger.info(f"Reporte generado: {output_path}")

    def _write_datos_sheet(self, wb):
        """Escribe hoja de datos consolidados para SUMIF en Resumen."""
        ws = wb.create_sheet("Datos")

        columnas = [
            ('Marca', 'Marca'),
            ('Mes_Año', 'Mes'),
            ('Total_Bonos_Sistema', 'Monto_Sistema'),
            ('Total_Pagado_Andes', 'Monto_Pagado'),
            ('Estado', 'Estado'),
        ]

        # Headers
        for col, (_, header) in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border

        # Datos
        for row_idx, row_data in enumerate(self.conciliacion, 2):
            for col, (col_orig, _) in enumerate(columnas, 1):
                value = row_data.get(col_orig, '')
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.thin_border
                if col in [3, 4]:
                    cell.number_format = self.currency_format

        datos_last_row = 1 + len(self.conciliacion)

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.freeze_panes = 'A2'

        return datos_last_row

    def _write_resumen_sheet(self, wb, datos_last_row):
        """Escribe hoja de resumen ejecutivo con fórmulas SUMIF/COUNTIF."""
        ws = wb.create_sheet("Resumen", 0)
        stats = self.statistics

        ws['A1'] = "CONCILIACIÓN DE COMISIONES - DANIEL ACHONDO"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        ws['A2'] = "Período: Diciembre 2024 - Septiembre 2025"
        ws['A3'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        marca_range = f"Datos!$A$2:$A${datos_last_row}"
        monto_sistema_range = f"Datos!$C$2:$C${datos_last_row}"
        monto_pagado_range = f"Datos!$D$2:$D${datos_last_row}"
        estado_range = f"Datos!$E$2:$E${datos_last_row}"

        # DESGLOSE POR MARCA
        row = 5
        ws[f'A{row}'] = "DESGLOSE POR MARCA"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        headers = ['Marca', 'Ventas', 'Con Bono', 'Con Pago', 'Pagados Exacto', 'Pendientes',
                   'Monto Sistema', 'Monto Pagado', 'Diferencia', '% Cobert.', '% Exacta', '% Monet.']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center')

        row += 1
        marcas = list(stats['by_marca'].keys())

        for marca in marcas:
            ws.cell(row=row, column=1, value=marca).border = self.thin_border
            ws.cell(row=row, column=2, value=f'=COUNTIF({marca_range},A{row})').border = self.thin_border
            ws.cell(row=row, column=3, value=f'=COUNTIFS({marca_range},A{row},{monto_sistema_range},">0")').border = self.thin_border
            ws.cell(row=row, column=4, value=f'=COUNTIFS({marca_range},A{row},{monto_pagado_range},">0")').border = self.thin_border
            ws.cell(row=row, column=5, value=f'=COUNTIFS({marca_range},A{row},{estado_range},"PAGADO")').border = self.thin_border
            ws.cell(row=row, column=6, value=f'=COUNTIFS({marca_range},A{row},{estado_range},"PENDIENTE")+COUNTIFS({marca_range},A{row},{estado_range},"PAGO PARCIAL")').border = self.thin_border

            cell_g = ws.cell(row=row, column=7, value=f'=SUMIF({marca_range},A{row},{monto_sistema_range})')
            cell_g.border = self.thin_border
            cell_g.number_format = self.currency_format

            cell_h = ws.cell(row=row, column=8, value=f'=SUMIF({marca_range},A{row},{monto_pagado_range})')
            cell_h.border = self.thin_border
            cell_h.number_format = self.currency_format

            cell_i = ws.cell(row=row, column=9, value=f'=G{row}-H{row}')
            cell_i.border = self.thin_border
            cell_i.number_format = self.currency_format

            ws.cell(row=row, column=10, value=f'=IF(C{row}>0,D{row}/C{row},0)').border = self.thin_border
            ws.cell(row=row, column=10).number_format = '0.0%'

            ws.cell(row=row, column=11, value=f'=IF(C{row}>0,E{row}/C{row},0)').border = self.thin_border
            ws.cell(row=row, column=11).number_format = '0.0%'

            ws.cell(row=row, column=12, value=f'=IF(G{row}>0,H{row}/G{row},0)').border = self.thin_border
            ws.cell(row=row, column=12).number_format = '0.0%'

            row += 1

        # Totales
        total_row = row
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        first_data_row = total_row - len(marcas)
        for col in range(2, 7):
            letter = get_column_letter(col)
            ws.cell(row=total_row, column=col,
                    value=f'=SUM({letter}{first_data_row}:{letter}{total_row-1})')
            ws.cell(row=total_row, column=col).border = self.thin_border
            ws.cell(row=total_row, column=col).font = Font(bold=True)

        for col in [7, 8, 9]:
            letter = get_column_letter(col)
            cell = ws.cell(row=total_row, column=col,
                           value=f'=SUM({letter}{first_data_row}:{letter}{total_row-1})')
            cell.border = self.thin_border
            cell.number_format = self.currency_format
            cell.font = Font(bold=True)

        ws.cell(row=total_row, column=10, value=f'=IF(C{total_row}>0,D{total_row}/C{total_row},0)')
        ws.cell(row=total_row, column=10).number_format = '0.0%'
        ws.cell(row=total_row, column=11, value=f'=IF(C{total_row}>0,E{total_row}/C{total_row},0)')
        ws.cell(row=total_row, column=11).number_format = '0.0%'
        ws.cell(row=total_row, column=12, value=f'=IF(G{total_row}>0,H{total_row}/G{total_row},0)')
        ws.cell(row=total_row, column=12).number_format = '0.0%'

        # Adjust column widths
        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _write_conciliacion_sheets_by_month(self, wb):
        """Escribe hojas de conciliación detallada por mes."""
        # Group by Mes_Año
        by_month = {}
        for r in self.conciliacion:
            mes = r.get('Mes_Año', 'Sin Fecha')
            if is_empty(mes):
                mes = 'Sin Fecha'
            by_month.setdefault(mes, []).append(r)

        # Sort by year + month
        def sort_key(item):
            rows = item[1]
            año = rows[0].get('Año', 9999) or 9999
            mes = rows[0].get('Mes_Num', 99) or 99
            return (año, mes)

        for mes_año, rows in sorted(by_month.items(), key=sort_key):
            sheet_name = mes_año[:31]  # Excel sheet name limit
            ws = wb.create_sheet(sheet_name)

            headers = ['VIN', 'Marca', 'Modelo', 'Estado',
                       'Bono Contado', 'Bono Crédito', 'Bono Volumen', 'Bono Créd. Extra',
                       'Total Sistema', 'Pagado Retail', 'Pagado Marca', 'Pagado CM',
                       'Pagado Flota', 'Total Pagado', 'Diferencia', 'Match']

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center')

            for row_idx, row_data in enumerate(rows, 2):
                values = [
                    row_data.get('VIN_NORM', ''),
                    row_data.get('Marca', ''),
                    row_data.get('Versión', ''),
                    row_data.get('Estado', ''),
                    row_data.get('Bono_Contado_Sistema', 0),
                    row_data.get('Bono_Credito_Sistema', 0),
                    row_data.get('Bono_Volumen_Sistema', 0),
                    row_data.get('Bono_Credito_Extra_Sistema', 0),
                    row_data.get('Total_Bonos_Sistema', 0),
                    row_data.get('Pagado_Bono_Retail', 0),
                    row_data.get('Pagado_Bono_Marca', 0),
                    row_data.get('Pagado_CM', 0),
                    row_data.get('Pagado_Flota', 0),
                    row_data.get('Total_Pagado_Andes', 0),
                    row_data.get('Diferencia', 0),
                    row_data.get('Match_Type', ''),
                ]

                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = self.thin_border
                    if col >= 5 and col <= 15:
                        cell.number_format = self.currency_format

                # Color by estado
                estado = row_data.get('Estado', '')
                if estado == 'PAGADO':
                    fill = self.fill_pagado
                elif estado == 'PENDIENTE':
                    fill = self.fill_pendiente
                elif estado == 'PAGO PARCIAL':
                    fill = self.fill_parcial
                else:
                    fill = None

                if fill:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col).fill = fill

            # Adjust widths
            ws.column_dimensions['A'].width = 20
            for col in range(2, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 14
            ws.freeze_panes = 'A2'

    def _write_pendientes_sheet(self, wb):
        """Escribe hoja de pendientes de pago."""
        ws = wb.create_sheet("Pendientes")

        pendientes = [r for r in self.conciliacion
                      if r.get('Estado') in ('PENDIENTE', 'PAGO PARCIAL')]

        headers = ['VIN', 'Marca', 'Modelo', 'Mes', 'Estado',
                   'Total Sistema', 'Total Pagado', 'Diferencia']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill_red
            cell.border = self.thin_border

        for row_idx, row_data in enumerate(pendientes, 2):
            values = [
                row_data.get('VIN_NORM', ''), row_data.get('Marca', ''),
                row_data.get('Versión', ''), row_data.get('Mes_Año', ''),
                row_data.get('Estado', ''),
                row_data.get('Total_Bonos_Sistema', 0),
                row_data.get('Total_Pagado_Andes', 0),
                row_data.get('Diferencia', 0),
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.thin_border
                if col >= 6:
                    cell.number_format = self.currency_format

            if row_data.get('Estado') == 'PENDIENTE':
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = self.fill_pendiente
            else:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = self.fill_parcial

        ws.column_dimensions['A'].width = 20
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.freeze_panes = 'A2'

    def _write_fuzzy_sheet(self, wb):
        """Escribe hoja de coincidencias difusas."""
        ws = wb.create_sheet("Coincidencias Difusas")

        headers = ['Marca', 'VIN Venta', 'VIN Pago', 'Similitud %',
                   'Modelo Venta', 'Modelo Pago', 'Monto Pago', 'GLOSA',
                   '¿Coincidencia Válida?', 'Comentarios']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill_orange
            cell.border = self.thin_border

        if not self.fuzzy_matches:
            ws.cell(row=2, column=1, value="No se encontraron coincidencias difusas")
            return

        sorted_matches = sorted(self.fuzzy_matches,
                                key=lambda m: (m.get('marca', ''), -m.get('similarity_score', 0)))

        for row_idx, match in enumerate(sorted_matches, 2):
            values = [
                match.get('marca', ''), match.get('vin_venta', ''),
                match.get('vin_pago', ''), match.get('similarity_score', 0),
                match.get('modelo_venta', ''), match.get('modelo_pago', ''),
                match.get('monto_pago', 0), match.get('glosa', ''),
                '', ''  # validation columns
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.thin_border

            # Color by score
            score = match.get('similarity_score', 0)
            if score >= 95:
                fill = PatternFill(start_color=COLORES['fuzzy_high'],
                                   end_color=COLORES['fuzzy_high'], fill_type='solid')
            elif score >= 90:
                fill = PatternFill(start_color=COLORES['fuzzy_medium'],
                                   end_color=COLORES['fuzzy_medium'], fill_type='solid')
            else:
                fill = PatternFill(start_color=COLORES['fuzzy_low'],
                                   end_color=COLORES['fuzzy_low'], fill_type='solid')

            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = fill

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        for col in range(4, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.freeze_panes = 'A2'

    def _write_pagos_sheet(self, wb):
        """Escribe hoja de detalle de pagos de Andes."""
        ws = wb.create_sheet("Detalle Pagos Andes")

        headers = ['VIN', 'MARCA', 'MODELO', 'GLOSA', 'MONTO', 'Tipo']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border

        for row_idx, pago in enumerate(self.pagos, 2):
            values = [
                pago.get('VIN_NORM', pago.get('vin', '')),
                pago.get('MARCA', ''), pago.get('MODELO', ''),
                pago.get('GLOSA', ''), pago.get('MONTO_NUM', pago.get('MONTO', 0)),
                pago.get('TIPO_GLOSA', '')
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.thin_border
                if col == 5:
                    cell.number_format = self.currency_format

        ws.column_dimensions['A'].width = 20
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.freeze_panes = 'A2'

    def _write_sin_venta_sheet(self, wb):
        """Escribe hoja de VINs sin venta registrada."""
        ws = wb.create_sheet("VINs Sin Venta")

        headers = ['VIN', 'MARCA', 'MODELO', 'GLOSA', 'MONTO']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill_red
            cell.border = self.thin_border

        if not self.unmatched_pagos:
            ws.cell(row=2, column=1, value="Todos los pagos tienen venta registrada")
            return

        for row_idx, pago in enumerate(self.unmatched_pagos, 2):
            values = [
                pago.get('VIN_NORM', pago.get('vin', '')),
                pago.get('MARCA', ''), pago.get('MODELO', ''),
                pago.get('GLOSA', ''), pago.get('MONTO_NUM', pago.get('MONTO', 0))
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.thin_border
                if col == 5:
                    cell.number_format = self.currency_format

        ws.column_dimensions['A'].width = 20
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.freeze_panes = 'A2'


def generate_conciliation_report(
    conciliacion: List[Dict],
    statistics: Dict[str, Any],
    pagos: List[Dict],
    unmatched_pagos: List[Dict],
    fuzzy_matches: List[Dict] = None,
    credit_status: List[Dict] = None,
    output_dir: str = "."
) -> str:
    """
    Genera reporte Excel de conciliación.

    Returns:
        Ruta al archivo generado
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Conciliacion_Comisiones_DA_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, filename)

    generator = ExcelReportGenerator(
        conciliacion=conciliacion,
        statistics=statistics,
        pagos=pagos,
        unmatched_pagos=unmatched_pagos,
        fuzzy_matches=fuzzy_matches,
        credit_status=credit_status
    )

    generator.generate_report(output_path)
    return output_path
